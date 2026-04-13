from __future__ import annotations

import csv
import json
from pathlib import Path
from zipfile import ZipFile
from uuid import uuid4
from xml.etree import ElementTree as ET

from docx import Document as DocxFile
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.entities import Artifact, Document
from app.services.storage import StorageService


class ArtifactService:
    def __init__(self) -> None:
        self.storage = StorageService()
        self.text_preview_char_limit = 24000
        self.table_preview_rows = 24
        self.table_preview_columns = 12
        self.pdf_preview_pages = 6
        self.presentation_preview_slides = 10

    async def create_source_file_artifact(
        self,
        session: AsyncSession,
        *,
        workspace_id,
        document_id,
        filename: str,
        mime_type: str,
        payload: bytes,
        metadata: dict | None = None,
    ) -> Artifact:
        suffix = Path(filename).suffix or ".bin"
        key = f"source-files/{workspace_id}/{document_id}/{uuid4().hex}{suffix}"
        self.storage.save_bytes(key, payload)
        artifact = Artifact(
            workspace_id=workspace_id,
            document_id=document_id,
            kind="source_file",
            title=filename,
            storage_key=key,
            metadata={"mime_type": mime_type, **(metadata or {})},
        )
        session.add(artifact)
        await session.flush()
        return artifact

    async def create_document_export(
        self,
        session: AsyncSession,
        *,
        document: Document,
        format: str = "markdown",
        include_metadata: bool = True,
        title: str | None = None,
    ) -> Artifact:
        rendered, extension, mime_type = self._render_document(document, format, include_metadata)
        artifact_title = title or f"{document.title}{extension}"
        key = f"exports/{document.workspace_id}/{document.id}/{uuid4().hex}{extension}"
        self.storage.save_text(key, rendered)
        artifact = Artifact(
            workspace_id=document.workspace_id,
            document_id=document.id,
            kind="document_export",
            title=artifact_title,
            storage_key=key,
            metadata={"format": format, "mime_type": mime_type},
        )
        session.add(artifact)
        await session.flush()
        return artifact

    async def list_artifacts(self, session: AsyncSession, *, workspace_id, document_id=None) -> list[Artifact]:
        statement = select(Artifact).where(Artifact.workspace_id == workspace_id)
        if document_id:
            statement = statement.where(Artifact.document_id == document_id)
        statement = statement.order_by(desc(Artifact.created_at))
        result = await session.execute(statement)
        return list(result.scalars().all())

    async def get_artifact(self, session: AsyncSession, artifact_id) -> Artifact | None:
        result = await session.execute(select(Artifact).where(Artifact.id == artifact_id))
        return result.scalar_one_or_none()

    def storage_path(self, key: str):
        return self.storage.resolve(key)

    def content_type(self, artifact: Artifact) -> str:
        return artifact.metadata_.get("mime_type") or self.storage.guess_content_type(artifact.storage_key)

    def build_preview(self, artifact: Artifact) -> dict:
        path = self.storage_path(artifact.storage_key)
        mime_type = self.content_type(artifact)
        extension = self._artifact_extension(artifact)
        preview_kind = self._preview_kind(mime_type, extension)
        warnings: list[str] = []
        preview = {
            "artifact_id": artifact.id,
            "workspace_id": artifact.workspace_id,
            "title": artifact.title,
            "kind": artifact.kind,
            "mime_type": mime_type,
            "preview_kind": preview_kind,
            "inline_supported": preview_kind in {"image", "audio", "video", "pdf", "html"},
            "text_content": None,
            "page_summaries": [],
            "table": None,
            "sheets": [],
            "slides": [],
            "metadata": artifact.metadata_ if isinstance(artifact.metadata_, dict) else {},
            "warnings": warnings,
            "size_bytes": path.stat().st_size if path.exists() else None,
        }

        if not path.exists():
            warnings.append("Artifact file is missing from storage.")
            preview["preview_kind"] = "missing"
            preview["inline_supported"] = False
            return preview

        try:
            if preview_kind in {"text", "markdown", "json", "html"}:
                preview["text_content"] = self._read_text_preview(path)
                return preview
            if preview_kind == "csv":
                preview["table"] = self._parse_csv_preview(path)
                return preview
            if preview_kind == "spreadsheet":
                preview["sheets"] = self._parse_spreadsheet_preview(path)
                return preview
            if preview_kind == "document":
                preview["text_content"] = self._parse_document_preview(path)
                return preview
            if preview_kind == "presentation":
                preview["slides"] = self._parse_presentation_preview(path)
                return preview
            if preview_kind == "pdf":
                preview["page_summaries"] = self._parse_pdf_preview(path)
                return preview
            return preview
        except Exception as exc:
            warnings.append(f"Preview extraction failed: {exc}")
            if preview["preview_kind"] not in {"image", "audio", "video", "pdf", "html"}:
                preview["preview_kind"] = "unsupported"
            return preview

    def _render_document(self, document: Document, format: str, include_metadata: bool) -> tuple[str, str, str]:
        normalized = format.lower()
        if normalized == "json":
            payload = {
                "title": document.title,
                "source_type": document.source_type,
                "source_uri": document.source_uri,
                "metadata": document.metadata_ if include_metadata else {},
                "content_text": document.content_text,
            }
            return json.dumps(payload, indent=2, ensure_ascii=False), ".json", "application/json"

        if normalized == "txt":
            prefix = ""
            if include_metadata:
                prefix = (
                    f"TITLE: {document.title}\nSOURCE TYPE: {document.source_type}\n"
                    f"SOURCE URI: {document.source_uri or '-'}\n\n"
                )
            return f"{prefix}{document.content_text}".strip(), ".txt", "text/plain"

        metadata_block = ""
        if include_metadata:
            metadata_lines = [
                f"- Source Type: {document.source_type}",
                f"- Source URI: {document.source_uri or '-'}",
            ]
            for key, value in document.metadata_.items():
                metadata_lines.append(f"- {key.replace('_', ' ').title()}: {value}")
            metadata_block = "\n".join(metadata_lines)

        rendered = f"# {document.title}\n\n"
        if metadata_block:
            rendered += f"{metadata_block}\n\n"
        rendered += document.content_text.strip()
        return rendered, ".md", "text/markdown"

    def _artifact_extension(self, artifact: Artifact) -> str:
        return Path(artifact.title or artifact.storage_key).suffix.lower()

    def _preview_kind(self, mime_type: str, extension: str) -> str:
        normalized_mime = (mime_type or "").lower()
        if normalized_mime.startswith("image/") or extension in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}:
            return "image"
        if normalized_mime.startswith("audio/") or extension in {".mp3", ".wav", ".ogg", ".m4a"}:
            return "audio"
        if normalized_mime.startswith("video/") or extension in {".mp4", ".webm", ".mov"}:
            return "video"
        if normalized_mime == "application/pdf" or extension == ".pdf":
            return "pdf"
        if "html" in normalized_mime or extension in {".html", ".htm"}:
            return "html"
        if normalized_mime == "application/json" or extension == ".json":
            return "json"
        if extension == ".csv" or "csv" in normalized_mime:
            return "csv"
        if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return "spreadsheet"
        if extension == ".docx":
            return "document"
        if extension == ".pptx":
            return "presentation"
        if extension in {".md", ".mdx"} or normalized_mime == "text/markdown":
            return "markdown"
        if normalized_mime.startswith("text/") or extension in {".txt", ".log", ".py", ".ts", ".tsx", ".js", ".jsx"}:
            return "text"
        return "unsupported"

    def _read_text_preview(self, path: Path) -> str:
        content = path.read_text(encoding="utf-8", errors="replace")
        return content[: self.text_preview_char_limit]

    def _parse_csv_preview(self, path: Path) -> dict:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            rows = [row for _, row in zip(range(self.table_preview_rows + 1), reader)]
        if not rows:
            return {"columns": [], "rows": []}
        columns = [str(value) for value in rows[0][: self.table_preview_columns]]
        body = [
            [str(value) for value in row[: self.table_preview_columns]]
            for row in rows[1 : self.table_preview_rows + 1]
        ]
        return {"columns": columns, "rows": body}

    def _parse_spreadsheet_preview(self, path: Path) -> list[dict]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        previews: list[dict] = []
        for sheet in workbook.worksheets[:4]:
            rows: list[list[str]] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index > self.table_preview_rows + 1:
                    break
                rows.append([
                    "" if value is None else str(value)
                    for value in row[: self.table_preview_columns]
                ])
            if rows:
                columns = rows[0]
                body = rows[1 : self.table_preview_rows + 1]
            else:
                columns = []
                body = []
            previews.append(
                {
                    "name": sheet.title,
                    "columns": columns,
                    "rows": body,
                }
            )
        workbook.close()
        return previews

    def _parse_document_preview(self, path: Path) -> str:
        document = DocxFile(path)
        paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
        return "\n\n".join(paragraphs)[: self.text_preview_char_limit]

    def _parse_pdf_preview(self, path: Path) -> list[str]:
        reader = PdfReader(str(path))
        summaries: list[str] = []
        for page_index, page in enumerate(reader.pages[: self.pdf_preview_pages], start=1):
            text = " ".join((page.extract_text() or "").split())
            summaries.append(f"Page {page_index}: {text[:420]}".strip())
        return summaries

    def _parse_presentation_preview(self, path: Path) -> list[dict]:
        namespace = {
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        }
        slides: list[dict] = []
        with ZipFile(path) as archive:
            slide_names = sorted(
                name
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            )
            for slide_number, slide_name in enumerate(slide_names[: self.presentation_preview_slides], start=1):
                xml_root = ET.fromstring(archive.read(slide_name))
                texts = [
                    text.strip()
                    for text in xml_root.findall(".//a:t", namespace)
                    if text.text and text.text.strip()
                ]
                slides.append(
                    {
                        "slide_number": slide_number,
                        "title": texts[0] if texts else None,
                        "bullets": texts[1:7] if len(texts) > 1 else [],
                    }
                )
        return slides
